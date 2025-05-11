import openpyxl
import os

os.chdir('Office') # changed current directory to Office folder by giving relative path

wb = openpyxl.load_workbook('example.xlsx')
print(type(wb)) # workbook object

print(wb.sheetnames)

sheet = wb['Sheet3'] # Worksheet

print(type(sheet))


print(sheet.title) # title of the sheet
 
print(wb.active.title) # show the active sheet's title

firstCellOfSheet3 = sheet['A1']
print(type(firstCellOfSheet3)) # Cell object

print(wb.active['A1'].value) # get first cell value of the active sheet
# since it was a date
print(type(wb.active['A1'].value)) # date 


wb = openpyxl.load_workbook('example.xlsx') # Get Workbook from current working directory
sheet = wb['Sheet1'] # Get Worksheet from the Workbook
print(sheet.title) # Get Worksheet title
c = sheet['B1'] # Get Cell from the Worksheet
print(c.value) # Apple
print('on') 
print(c.row, 'row') # 1
print(c.column, 'column') # 2
print('and it is the cell', c.coordinate) # B1 - kinda Cell title 

# get cells by row and column
print(sheet.cell(row=1, column=2).value) # Apple
# or simply 
print(sheet.cell(1,2).value) # Apple