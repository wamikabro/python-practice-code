import openpyxl, os

os.chdir('Office')

wb = openpyxl.load_workbook('example.xlsx') # workbook

sheet1 = wb['Sheet1']

print(sheet1.cell(1,2).value)

for i in range(1,8): # 1 - 7
    print(i, sheet1.cell(i,2).value) # row 1 (print column 2)
                                     # row 3 (print column 2) ...

print(sheet1.max_row) # how many rows, highest number of rows 
print(sheet1.max_column) # how many columns, highest number of columns